"""Knowledge-node Excel parser (W5a).

接受「整合-知識節點大全」xlsx；第一個 sheet「整合」。
欄位（依使用者提供格式）：
    A 序號 / B 星空圖名稱 / C 學習階段（2=middle、3=upper）/
    D 學習內容編碼（parent_code）/ E 學習內容名稱（parent_name）/
    F 小節點編碼（id）/ G 小節點名稱（name）/
    H 影片名稱（video_title）/ I 教學影片網址（video_url）
"""
from __future__ import annotations

from typing import BinaryIO

from openpyxl import load_workbook

STAGE_TO_BAND = {1: "lower", 2: "middle", 3: "upper"}


class NodeExcelParseError(ValueError):
    """Raised when the uploaded Excel does not match the expected format."""


def _norm(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_nodes_xlsx(file: BinaryIO, sheet_name: str = "整合") -> list[dict]:
    """Parse the integrated-nodes xlsx.

    Returns list of {id, name, parent_code, parent_name, grade_band, video_title, video_url}.
    Raises NodeExcelParseError on structural problems.
    """
    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception as exc:
        raise NodeExcelParseError(f"INVALID_XLSX: {exc}") from exc

    if sheet_name not in wb.sheetnames:
        raise NodeExcelParseError(f"SHEET_NOT_FOUND:{sheet_name}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise NodeExcelParseError("EMPTY_SHEET")

    # Detect header: first row's column F (index 5) header should be 「小節點編碼」 or first column should be a header.
    first = rows[0]
    looks_like_header = False
    if len(first) > 5 and isinstance(first[5], str) and not first[5].strip().isdigit():
        # If column F's first cell isn't a node id like "INa-Ⅱ-1-01", treat as header
        looks_like_header = "-" not in (first[5] or "")
    if looks_like_header:
        rows = rows[1:]

    seen_ids: set[str] = set()
    parsed: list[dict] = []
    for idx, raw in enumerate(rows, start=1):
        # Skip fully empty rows
        if not any(_norm(c) for c in raw):
            continue
        if len(raw) < 9:
            raise NodeExcelParseError(f"ROW_{idx}_MISSING_COLUMNS")
        stage_raw = _norm(raw[2])
        parent_code = _norm(raw[3]) or None
        parent_name = _norm(raw[4]) or None
        node_id = _norm(raw[5])
        node_name = _norm(raw[6])
        video_title = _norm(raw[7]) or None
        video_url = _norm(raw[8]) or None

        if not node_id:
            raise NodeExcelParseError(f"ROW_{idx}_MISSING_ID")
        if not node_name:
            raise NodeExcelParseError(f"ROW_{idx}_MISSING_NAME")
        if node_id in seen_ids:
            raise NodeExcelParseError(f"DUPLICATE_NODE_ID:{node_id}")
        seen_ids.add(node_id)

        try:
            stage = int(stage_raw) if stage_raw else 0
        except ValueError as exc:
            raise NodeExcelParseError(f"ROW_{idx}_INVALID_STAGE:{stage_raw}") from exc
        grade_band = STAGE_TO_BAND.get(stage)
        if grade_band is None:
            raise NodeExcelParseError(f"ROW_{idx}_INVALID_STAGE:{stage}")

        if len(node_id) > 64:
            raise NodeExcelParseError(f"ROW_{idx}_ID_TOO_LONG")
        if len(node_name) > 256:
            raise NodeExcelParseError(f"ROW_{idx}_NAME_TOO_LONG")

        parsed.append({
            "id": node_id,
            "name": node_name,
            "parent_code": parent_code,
            "parent_name": parent_name,
            "grade_band": grade_band,
            "video_title": video_title,
            "video_url": video_url,
        })

    if not parsed:
        raise NodeExcelParseError("NO_VALID_ROWS")
    return parsed
