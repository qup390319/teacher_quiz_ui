"""Excel parser for class roster import (W3).

Spec：
- 接受 .xlsx 檔案；第一個 sheet
- 兩欄必填：座號 (seat) / 姓名 (name)
- 允許表頭（中文標題列），自動偵測並略過
- 座號需為正整數；姓名需為非空字串

回傳 list of {seat, name}，已驗證 + 去重（同座號回 ValueError）。
"""
from __future__ import annotations

from typing import BinaryIO

from openpyxl import load_workbook


class ExcelParseError(ValueError):
    """Raised when the uploaded Excel does not match the expected format."""


def _norm_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _looks_like_header(row: tuple) -> bool:
    """If both cells are non-numeric strings, treat as header."""
    if len(row) < 2:
        return False
    a, b = _norm_cell(row[0]), _norm_cell(row[1])
    if not a or not b:
        return False
    # Header heuristic: seat column is not parseable as a positive integer.
    try:
        int(a)
        return False
    except ValueError:
        return True


def parse_roster_xlsx(file: BinaryIO) -> list[dict]:
    """Parse a roster .xlsx file into a list of {seat:int, name:str}.

    Raises ExcelParseError on any structural problem (missing column, bad seat,
    duplicate seat, empty name, etc.).
    """
    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises various subclasses
        raise ExcelParseError(f"INVALID_XLSX: {exc}") from exc

    ws = wb.active
    if ws is None:
        raise ExcelParseError("EMPTY_WORKBOOK")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ExcelParseError("EMPTY_SHEET")

    # Skip header if present
    if _looks_like_header(rows[0]):
        rows = rows[1:]

    seats_seen: set[int] = set()
    parsed: list[dict] = []
    for idx, raw in enumerate(rows, start=1):
        seat_raw = _norm_cell(raw[0] if len(raw) > 0 else "")
        name_raw = _norm_cell(raw[1] if len(raw) > 1 else "")
        # Skip fully-empty trailing rows
        if not seat_raw and not name_raw:
            continue
        if not seat_raw:
            raise ExcelParseError(f"ROW_{idx}_MISSING_SEAT")
        if not name_raw:
            raise ExcelParseError(f"ROW_{idx}_MISSING_NAME")
        try:
            seat = int(seat_raw)
        except ValueError as exc:
            raise ExcelParseError(f"ROW_{idx}_INVALID_SEAT:{seat_raw}") from exc
        if seat <= 0:
            raise ExcelParseError(f"ROW_{idx}_INVALID_SEAT:{seat}")
        if len(name_raw) > 64:
            raise ExcelParseError(f"ROW_{idx}_NAME_TOO_LONG")
        if seat in seats_seen:
            raise ExcelParseError(f"DUPLICATE_SEAT:{seat}")
        seats_seen.add(seat)
        parsed.append({"seat": seat, "name": name_raw})

    if not parsed:
        raise ExcelParseError("NO_VALID_ROWS")

    return parsed
